import openpyxl
import os
IP_file=openpyxl.load_workbook("dai_IP.xlsx")
IP_list=IP_file["Sheet1"]
dict_IP={}
for row_ip in range(2,IP_list.max_row +1):
    servername=IP_list.cell(row_ip,2).value
    IP_per_server=IP_list.cell(row_ip,3).value
    Status=IP_list.cell(row_ip,4)
    dict_IP[servername]=IP_per_server
#    print("Các IP là:",IP_per_server)

    #Check ping.
    host_IP=IP_per_server
    response=os.popen(f"ping {host_IP} -n 2").read()

    if "Request timed out" in response:
        Show="is DOWN!"
        print(host_IP,"is DOWN!")
        Status.value=Show
    elif "Destination host unreachable" in response:
        Show ="is DOWN!"
        print(host_IP,"is DOWN!")
        Status.value = Show
    elif "Ping request could not find host None" in response:
        Show = "could not find host None!"
        print(host_IP, "could not find host None!")
        Status.value = Show
#        print(response)
    else:
        Show ="is UP!"
        print(host_IP,"is UP!")
        Status.value = Show

IP_file.save("IP_status_Version3.xlsx")
#print(dict_IP)