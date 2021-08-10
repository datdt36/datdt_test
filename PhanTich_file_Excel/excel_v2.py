import openpyxl
inv_file=openpyxl.load_workbook("inventory.xlsx")
product_list=inv_file["Sheet1"]
product_per_supplier={}
total_value_per_supplier={}
List_inventory_less10=""
List_less10_dict={}
# print(product_list.cell(2,4).value)
for product_row in range(2,product_list.max_row+1):
    product_no = product_list.cell(product_row, 1).value
    supplier_name=product_list.cell(product_row,4).value
    product_inventory_price=product_list.cell(product_row,5)
    invetory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_inventory_price.value=invetory*price


#    print("Tiền tồn kho của sản phẩm",product_no,"là: ",product_inventory_price.value)


    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name]+=1
    else:
        product_per_supplier[supplier_name]=1

    #tổng giá trị tồn kho của mỗi công ty.
    # invetory=product_list.cell(product_row,2).value
    # price=product_list.cell(product_row,3).value

    if supplier_name in total_value_per_supplier:
        curent_total=invetory*price
        total_value_per_supplier[supplier_name]=total_value_per_supplier[supplier_name]+curent_total
    else:
        total_value_per_supplier[supplier_name]=invetory*price

    #Print product no is nho hon 10.
    if invetory<10:
        List_inventory_less10=List_inventory_less10+str(product_no)+"; "
        List_less10_dict[product_no]=product_list.cell(product_row,2).value


print("Các sản phẩm có số hàng tồn dưới 10 là:",List_inventory_less10.strip("; "))
print("Các sản phẩm có số hàng tồn dưới 10 là:",List_less10_dict)
print("Số lượng sản phẩm của mỗi cty là:",product_per_supplier)
print("Số tiền tính theo hàng tồn của mỗi cty là:",total_value_per_supplier)

#inv_file.save("maynguqua4.xlsx")



